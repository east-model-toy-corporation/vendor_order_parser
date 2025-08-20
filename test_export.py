import os
import pandas as pd
from data_processor import generate_erp_excel

# Simple logger
def logger(msg):
    print(msg)

# Build sample all_products
all_products = []

global_info = {'寄件廠商': '示例廠商', '結單日期': '2025/12/15'}
product1 = {
    '國際條碼': '1234567890123',
    '貨號': 'SKU123',
    '品名': 'POP RACE | 1/64 保時捷 SINGER 964 TIFFANY BLUE',
    '起始進價': '100',
    '建議售價': '200',
    '備註': '測試備註',
    '預計發售月份': '2025-11-01'
}
product2 = {
    '國際條碼': '9876543210987',
    '貨號': 'SKU456',
    '品名': 'OTHERBRAND | 測試商品 2',
    '起始進價': '150',
    '建議售價': '300',
    '備註': '備註2',
    '預計發售月份': '2026.01'
}
all_products.append({'global_info': global_info, 'product_data': product1})
all_products.append({'global_info': global_info, 'product_data': product2})

output_path = os.path.join(os.path.dirname(__file__), 'test_output.xlsx')

# Generate the excel
generate_erp_excel(all_products, output_path, logger)

# Inspect the created file for formulas in M3 and Q3 using openpyxl
try:
    from openpyxl import load_workbook
    wb = load_workbook(output_path, data_only=False)
    ws = wb['ERP']
    def cell_info(cell_ref):
        cell = ws[cell_ref]
        return {'value': cell.value, 'formula': cell.value if cell.data_type=='f' else None, 'data_type': cell.data_type}

    m3 = ws['M3']
    q3 = ws['Q3']
    print('\nM3 cell (brand) raw:', m3.value, 'data_type=', m3.data_type)
    print('Q3 cell (vendor) raw:', q3.value, 'data_type=', q3.data_type)
except Exception as e:
    print('openpyxl inspect failed:', e)

# Also read back with pandas to show the values as loaded (formulas may be blank if not evaluated)
try:
    df = pd.read_excel(output_path, sheet_name='ERP')
    print('\nPandas read head:')
    print(df.head(5).to_string(index=False))
except Exception as e:
    print('pandas read failed:', e)

print('\nTest export finished, output file:', output_path)
